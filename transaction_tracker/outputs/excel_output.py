# transaction_tracker/outputs/excel_output.py

from datetime import datetime
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, numbers

from transaction_tracker.outputs.base import BaseOutput
from transaction_tracker.core.categorizer import categorize


class ExcelOutput(BaseOutput):
    """Local Excel workbook similar to SheetsOutput."""

    MONTH_FMT = "%B %Y"
    ALL_DATA = "AllData"
    SUMMARY = "Summary"

    def __init__(self, config):
        self.config = config
        self.output_dir = config.get('output_dir', 'data')
        os.makedirs(self.output_dir, exist_ok=True)

    def append(self, transactions):
        if not transactions:
            print("No transactions to write.")
            return

        months = sorted({tx.date.strftime('%Y-%m') for tx in transactions})
        first_dt = datetime.strptime(months[0], "%Y-%m")
        year = first_dt.year
        out_path = os.path.join(self.output_dir, f"Budget{year}.xlsx")

        wb = Workbook()
        if wb.active.title == "Sheet":
            wb.remove(wb.active)

        # Monthly tabs
        all_rows = []
        for month_str in months:
            dt = datetime.strptime(month_str, "%Y-%m")
            tab_title = dt.strftime(self.MONTH_FMT)
            ws = wb.create_sheet(title=tab_title)
            ws.append(["date", "description", "merchant", "category", "amount"])
            ws.freeze_panes = "A2"
            month_rows = []
            for tx in transactions:
                if tx.date.strftime("%Y-%m") != month_str:
                    continue
                cat = categorize(tx, self.config["categories"]) or ""
                row = [
                    tx.date.isoformat(),
                    tx.description,
                    tx.merchant,
                    cat,
                    float(tx.amount),
                ]
                ws.append(row)
                month_rows.append(row)
                all_rows.append([tab_title] + row)
            # Pivot per category on the right
            if month_rows:
                df_month = pd.DataFrame(month_rows, columns=["date", "description", "merchant", "category", "amount"])
                pivot = (
                    df_month.pivot_table(index="category", values="amount", aggfunc="sum")
                    .reset_index()
                    .sort_values("category")
                )
                start_col = 7  # column G
                ws.cell(row=1, column=start_col, value="category")
                ws.cell(row=1, column=start_col + 1, value="amount")
                for idx, r in pivot.iterrows():
                    ws.cell(row=2 + idx, column=start_col, value=r["category"])
                    ws.cell(row=2 + idx, column=start_col + 1, value=float(r["amount"] or 0))
                self._format_amount_column(ws, start_col + 1)
            self._format_amount_column(ws, 5)

        # AllData tab
        all_ws = wb.create_sheet(title=self.ALL_DATA)
        all_ws.append(["month", "date", "description", "merchant", "category", "amount"])
        for row in all_rows:
            all_ws.append(row)
        all_ws.freeze_panes = "A2"
        self._format_amount_column(all_ws, 6)

        # Summary tab with month columns and categories rows
        sum_ws = wb.create_sheet(title=self.SUMMARY)
        df = pd.DataFrame(
            all_rows,
            columns=["month", "date", "description", "merchant", "category", "amount"],
        )
        pivot = df.pivot_table(
            index="category",
            columns="month",
            values="amount",
            aggfunc="sum",
            fill_value=0,
        )
        # Ensure columns in chronological order
        month_titles = [datetime.strptime(m, "%Y-%m").strftime(self.MONTH_FMT) for m in months]
        pivot = pivot.reindex(columns=month_titles, fill_value=0).sort_index()
        header = ["category"] + list(pivot.columns)
        sum_ws.append(header)
        for cat in pivot.index:
            sum_ws.append([cat] + [float(pivot.at[cat, m]) for m in pivot.columns])
        sum_ws.freeze_panes = "B2"
        for idx in range(2, len(header) + 1):
            self._format_amount_column(sum_ws, idx)

        wb.save(out_path)
        print(f"Written Excel workbook {out_path}")

    def _format_amount_column(self, ws, idx):
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E5E5E5", end_color="E5E5E5", fill_type="solid")
        col_letter = get_column_letter(idx)
        for cell in ws[col_letter][1:]:
            cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
