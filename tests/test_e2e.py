import os
import csv
import sys
import types
import yaml
from click.testing import CliRunner
import openpyxl
import xlsxwriter
import zipfile
import pytest

# Provide minimal google modules so sheets_output can be imported without
# installing heavy dependencies.
fake_gspread = types.ModuleType("gspread")
fake_gspread.exceptions = types.SimpleNamespace(
    SpreadsheetNotFound=Exception, WorksheetNotFound=Exception
)
fake_gspread.authorize = lambda *_a, **_k: None
fake_gspread.__path__ = []
def _rowcol_to_a1(row, col):
    letters = ""
    while col:
        col, rem = divmod(col - 1, 26)
        letters = chr(rem + ord('A')) + letters
    return f"{letters}{row}"
fake_utils = types.ModuleType("gspread.utils")
fake_utils.rowcol_to_a1 = _rowcol_to_a1
fake_gspread.utils = fake_utils

fake_service_account = types.ModuleType("google.oauth2.service_account")

class FakeCreds:
    @classmethod
    def from_service_account_file(cls, *_a, **_k):
        return cls()

fake_service_account.Credentials = FakeCreds

fake_discovery = types.ModuleType("googleapiclient.discovery")

def fake_build(*_a, **_k):
    class Dummy:
        def __getattr__(self, name):
            return lambda *a, **k: self

    return Dummy()

fake_discovery.build = fake_build

sys.modules.setdefault("gspread", fake_gspread)
sys.modules.setdefault("gspread.utils", fake_utils)
sys.modules.setdefault("google.oauth2.service_account", fake_service_account)
sys.modules.setdefault("googleapiclient.discovery", fake_discovery)

import transaction_tracker.ai as ai


class DummyProvider:
    def generate(self, messages):
        return "AI Report"


ai.get_provider_from_env = lambda: DummyProvider()

from transaction_tracker.cli import main as cli
from transaction_tracker.outputs import sheets_output


def write_config(tmp_path, data_dir):
    cfg = {
        'bank_loaders': {
            'tdvisa': 'transaction_tracker.loaders.tdvisa.TDVisaLoader',
        },
        'output_modules': {
            'csv': 'transaction_tracker.outputs.csv_output.CSVOutput',
            'sheets': 'transaction_tracker.outputs.sheets_output.SheetsOutput',
            'excel': 'transaction_tracker.outputs.excel_output.ExcelOutput',
        },
        'categories': {
            'restaurants': ['restaurant'],
            'groceries': ['grocery'],
        },
        'output_dir': str(data_dir),
        'google': {
            'service_account_file': 'creds.json',
            'sheet_folder_id': 'folder',
            'owner_email': 'owner@example.com',
        },
    }
    path = tmp_path / 'config.yaml'
    with open(path, 'w') as f:
        yaml.safe_dump(cfg, f)
    return path


def write_tdvisa_sample(path):
    rows = [
        ['05/02/2025', 'Grocery Store', '56.78', '', '1000'],
        ['05/03/2025', 'Restaurant A', '12.34', '', '990'],
    ]
    with open(path, 'w', newline='') as f:
        csv.writer(f).writerows(rows)


def write_manual(path):
    path.write_text(
        """\
- date: 2025-05-04
  description: Farmers Market
  merchant: CASH
  amount: 10
"""
    )


def test_cli_csv_output(tmp_path):
    stmts = tmp_path / 'stmts'
    stmts.mkdir()
    td_file = stmts / 'tdvisa.csv'
    write_tdvisa_sample(td_file)
    manual = tmp_path / 'manual.yaml'
    write_manual(manual)
    cfg_path = write_config(tmp_path, tmp_path / 'data')

    runner = CliRunner()
    res = runner.invoke(
        cli,
        ['--dir', str(stmts), '--output', 'csv', '--config', str(cfg_path), '--manual-file', str(manual)]
    )
    assert res.exit_code == 0, res.output
    out_csv = tmp_path / 'data' / 'Budget2025.csv'
    assert out_csv.exists()
    with open(out_csv) as f:
        lines = [l.strip() for l in f]
    assert lines[0] == 'date,description,merchant,category,amount'
    assert len(lines) == 4  # header + 3 rows
    assert any('restaurants' in l for l in lines[1:])
    assert any('groceries' in l for l in lines[1:])


def test_cli_excel_output(tmp_path):
    stmts = tmp_path / 'stmts'
    stmts.mkdir()
    td_file = stmts / 'tdvisa.csv'
    write_tdvisa_sample(td_file)
    manual = tmp_path / 'manual.yaml'
    write_manual(manual)
    cfg_path = write_config(tmp_path, tmp_path / 'data')

    runner = CliRunner()
    res = runner.invoke(
        cli,
        ['--dir', str(stmts), '--output', 'excel', '--config', str(cfg_path), '--manual-file', str(manual)]
    )
    assert res.exit_code == 0, res.output
    out_xlsx = tmp_path / 'data' / 'Budget2025.xlsx'
    assert out_xlsx.exists()
    wb = openpyxl.load_workbook(out_xlsx, data_only=True)
    assert 'May 2025' in wb.sheetnames
    assert 'AllData' in wb.sheetnames
    assert 'Summary' in wb.sheetnames

    may_ws = wb['May 2025']
    assert may_ws['A1'].value == 'date'
    assert may_ws['E1'].value == 'amount'
    amounts = [c[0].value for c in may_ws.iter_rows(min_row=2, min_col=5, max_col=5)]
    assert amounts == sorted(amounts, key=abs, reverse=True)

    # Ensure PivotTables were created in the workbook
    if hasattr(xlsxwriter.Workbook, 'add_pivot_table'):
        with zipfile.ZipFile(out_xlsx) as zf:
            pivot_xml = ''
            pivot_files = [
                n for n in zf.namelist() if n.startswith('xl/pivotTables/pivotTable')
            ]
            assert len(pivot_files) == 1
            for name in pivot_files:
                pivot_xml += zf.read(name).decode('utf-8')

            import xml.etree.ElementTree as ET

            wb_tree = ET.fromstring(zf.read('xl/workbook.xml'))
            rels_tree = ET.fromstring(zf.read('xl/_rels/workbook.xml.rels'))
            ns_main = {
                'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
            }
            ns_rel = {
                'rel': 'http://schemas.openxmlformats.org/package/2006/relationships'
            }
            r_id = None
            for sheet in wb_tree.find('main:sheets', ns_main).findall('main:sheet', ns_main):
                if sheet.get('name') == 'Summary':
                    r_id = sheet.get(
                        '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id'
                    )
                    break
            assert r_id is not None
            target = None
            for rel in rels_tree.findall('rel:Relationship', ns_rel):
                if rel.get('Id') == r_id:
                    target = rel.get('Target')
                    break
            assert target is not None
            summary_rel_path = f"xl/worksheets/_rels/{os.path.basename(target)}.rels"
            summary_rel_xml = zf.read(summary_rel_path).decode('utf-8')
            assert 'pivotTable' not in summary_rel_xml

        assert 'name="Pivot_May_2025"' in pivot_xml
        assert 'Pivot_Summary' not in pivot_xml

    summary_ws = wb['Summary']
    rows = list(summary_ws.values)
    assert ('May 2025 Total', None, pytest.approx(79.12)) in rows
    assert rows[-1][0] == 'Grand Total'
    assert rows[-1][2] == pytest.approx(79.12)


class FakeWorksheet:
    def __init__(self, title):
        self.title = title
        self.rows = []
        self.spreadsheet = types.SimpleNamespace(id='123')

    def update_title(self, title):
        self.title = title

    def clear(self):
        self.rows = []

    def update(self, *_args, **_kwargs):
        if _args:
            self.rows = _args[1]

    def set_basic_filter(self, *_a, **_k):
        pass

    def sort(self, *_a, **_k):
        pass

    def get_all_values(self):
        # Simulate pivot-table columns appended to the right of transaction
        # data, which happens in the real Google Sheets output. Only
        # non-aggregate tabs (i.e., monthly tabs) include these extra values.
        if self.title not in ('AllData', 'Summary', 'Charts'):
            return [row + ['0'] for row in self.rows]
        return self.rows


class FakeSpreadsheet:
    def __init__(self):
        self.id = '123'
        self.sheet1 = FakeWorksheet('Sheet1')
        self._worksheets = [self.sheet1]

    def worksheet(self, title):
        for ws in self._worksheets:
            if ws.title == title:
                return ws
        raise sheets_output.gspread.exceptions.WorksheetNotFound

    def add_worksheet(self, title, rows='100', cols='10'):
        ws = FakeWorksheet(title)
        self._worksheets.append(ws)
        return ws

    def worksheets(self):
        return self._worksheets

    def share(self, *_a, **_k):
        pass

    def list_permissions(self):
        return [{'emailAddress': 'owner@example.com', 'role': 'writer'}]


def setup_sheet_mocks(monkeypatch):
    class FakeCreds:
        @classmethod
        def from_service_account_file(cls, *_a, **_k):
            return cls()

    class FakeClient:
        def __init__(self):
            self.sheet = FakeSpreadsheet()

        def open(self, *_a, **_k):
            raise sheets_output.gspread.exceptions.SpreadsheetNotFound

        def create(self, *_a, **_k):
            return self.sheet

    class FakeExceptions:
        class SpreadsheetNotFound(Exception):
            pass

        class WorksheetNotFound(Exception):
            pass

    monkeypatch.setattr(sheets_output, 'Credentials', FakeCreds)
    client = FakeClient()
    monkeypatch.setattr(
        sheets_output,
        'gspread',
        types.SimpleNamespace(authorize=lambda *_a, **_k: client, exceptions=FakeExceptions)
    )
    def fake_build(*_a, **_k):
        sheet = client.sheet

        class Dummy:
            def files(self):
                return self
            def get(self, *a, **k):
                fields = k.get('fields', '')
                if 'parents' in fields:
                    data = {'parents': []}
                elif 'sheets.properties' in fields:
                    data = {
                        'sheets': [
                            {'properties': {'title': 'Sheet1', 'sheetId': 1}},
                            {'properties': {'title': 'May 2025', 'sheetId': 2}},
                            {'properties': {'title': 'AllData', 'sheetId': 3}},
                            {'properties': {'title': 'Summary', 'sheetId': 4}},
                            {'properties': {'title': 'Charts', 'sheetId': 5}},
                        ]
                    }
                else:
                    data = {}
                return types.SimpleNamespace(execute=lambda: data)
            def update(self, *a, **k):
                return types.SimpleNamespace(execute=lambda: {})
            def spreadsheets(self):
                return self
            def batchUpdate(self, *a, **k):
                return types.SimpleNamespace(execute=lambda: {})
            def values(self):
                class ValuesDummy:
                    def batchClear(self_inner, *a, **k):
                        ranges = k.get('body', {}).get('ranges', [])
                        for rng in ranges:
                            title = rng.split('!')[0].strip("'")
                            try:
                                ws = sheet.worksheet(title)
                            except sheets_output.gspread.exceptions.WorksheetNotFound:
                                continue
                            ws.rows = []
                        return types.SimpleNamespace(execute=lambda: {})
                    def batchUpdate(self_inner, *a, **k):
                        data = k.get('body', {}).get('data', [])
                        for item in data:
                            rng = item.get('range', '')
                            title = rng.split('!')[0].strip("'")
                            try:
                                ws = sheet.worksheet(title)
                            except sheets_output.gspread.exceptions.WorksheetNotFound:
                                continue
                            ws.rows = item.get('values', [])
                        return types.SimpleNamespace(execute=lambda: {})
                return ValuesDummy()
        return Dummy()

    monkeypatch.setattr(sheets_output, 'build', fake_build)
    return client


def test_cli_sheets_output(tmp_path, monkeypatch):
    setup_sheet_mocks(monkeypatch)
    stmts = tmp_path / 'stmts'
    stmts.mkdir()
    td_file = stmts / 'tdvisa.csv'
    write_tdvisa_sample(td_file)
    manual = tmp_path / 'manual.yaml'
    write_manual(manual)
    cfg_path = write_config(tmp_path, tmp_path / 'data')

    runner = CliRunner()
    res = runner.invoke(
        cli,
        ['--dir', str(stmts), '--output', 'sheets', '--config', str(cfg_path), '--manual-file', str(manual)]
    )
    assert res.exit_code == 0, res.output
    assert 'Appended 3 transaction' in res.output


def test_all_data_preserves_amounts(tmp_path, monkeypatch):
    client = setup_sheet_mocks(monkeypatch)
    stmts = tmp_path / 'stmts'
    stmts.mkdir()
    td_file = stmts / 'tdvisa.csv'
    write_tdvisa_sample(td_file)
    manual = tmp_path / 'manual.yaml'
    write_manual(manual)
    cfg_path = write_config(tmp_path, tmp_path / 'data')

    runner = CliRunner()
    res = runner.invoke(
        cli,
        ['--dir', str(stmts), '--output', 'sheets', '--config', str(cfg_path), '--manual-file', str(manual)]
    )
    assert res.exit_code == 0, res.output
    all_rows = client.sheet.worksheet('AllData').rows
    amounts = sorted(r[-1] for r in all_rows[1:])
    assert amounts == sorted([56.78, 12.34, 10.0])


def test_cli_ai_report(tmp_path, monkeypatch):
    stmts = tmp_path / 'stmts'
    stmts.mkdir()
    td_file = stmts / 'tdvisa.csv'
    write_tdvisa_sample(td_file)
    manual = tmp_path / 'manual.yaml'
    write_manual(manual)
    cfg_path = write_config(tmp_path, tmp_path / 'data')

    env_file = tmp_path / '.env'
    env_file.write_text('BUDGIFY_LLM_PROVIDER=huggingface\nHF_API_TOKEN=dummy\n')
    runner = CliRunner()
    res = runner.invoke(
        cli,
        ['--dir', str(stmts), '--output', 'csv', '--config', str(cfg_path), '--manual-file', str(manual), '--env-file', str(env_file), '--ai-report']
    )
    assert res.exit_code == 0, res.output
    assert 'AI Report:' in res.output
